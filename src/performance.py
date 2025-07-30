import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font

import pandas as pd
import warnings
import config
import numpy as np

BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Folder where performance.py is located
output_file = os.path.join(BASE_DIR, "..", "results", "Performance.xlsx")
warnings.filterwarnings("ignore", category=pd.errors.PerformanceWarning)

projects = config.PROJECTS
runs = config.RUN
rounds = 2

MODULES_PROJECTS = {k: v for k, v in config.MODULES_PROJECTS.items() if k != "roller"}
init_label = "Init"
no_mod_label = "No Modular"
no_dep_label = "No Dependency"

cpuResults = []
memoryResults = []
recordResults = []


def load_csv_comparison(name, module, number, base_dir):
    if module:
        base_path = os.path.join(base_dir, "..", "performance_data", name, module)
    else:
        base_path = os.path.join(base_dir, "..", "performance_data", name)
    if not os.path.exists(base_path):
        print(f"⚠️ Folder not found: {base_path}. Skipping loading.")
        return None, None, None, None, None, None, None, None, None

    # CPU
    df_init = pd.read_csv(os.path.join(base_path, f"InitCPU_Load_telemetry{number}.csv"))
    df_no_mod = pd.read_csv(os.path.join(base_path, f"NoModularCPU_Load_telemetry{number}.csv"))
    df_no_dep = pd.read_csv(os.path.join(base_path, f"NoDependencyCPU_Load_telemetry{number}.csv"))

    # Memory
    df_memory_init = pd.read_csv(os.path.join(base_path, f'InitMemory_telemetry{number}.csv'))
    df_memory_no_mod = pd.read_csv(os.path.join(base_path, f'NoModularMemory_telemetry{number}.csv'))
    df_memory_no_dep = pd.read_csv(os.path.join(base_path, f'NoDependencyMemory_telemetry{number}.csv'))

    # Recorded Objects
    df_record_init = pd.read_csv(os.path.join(base_path, f'InitRecorded_Objects{number}.csv'))
    df_record_no_mod = pd.read_csv(os.path.join(base_path, f'NoModularRecorded_Objects{number}.csv'))
    df_record_no_dep = pd.read_csv(os.path.join(base_path, f'NoDependencyRecorded_Objects{number}.csv'))

    return (
        df_init, df_no_mod, df_no_dep,              # CPU
        df_memory_init, df_memory_no_mod, df_memory_no_dep,  # Memory
        df_record_init, df_record_no_mod, df_record_no_dep   # Recorded
    )


def get_stats(df, label_prefix, rounds):
    stats = []
    for col in df.columns:
        if col.lower() != 'time [s]':  # exclude the time column
            stats.append({
                'Performance': f'{col} {label_prefix}',
                'Max': round(df[col].max(), rounds),
                'Min': round(df[col].min(), rounds),
                'Var': round(df[col].var(), rounds),
                'AVG': round(df[col].mean(), rounds)
            })
    return pd.DataFrame(stats)


def get_stats_memory(df, label_prefix, rounds):
    stats = []
    for col in df.columns:
        if col.lower() != 'time [s]':  # exclude the time column
            max_val = df[col].max() / (1024 ** 2)
            min_val = df[col].min() / (1024 ** 2)
            var_val = df[col].var() / (1024 ** 2)
            avg_val = df[col].mean() / (1024 ** 2)
            stats.append({
                'Performance': f'{col} {label_prefix}',
                'Max': round(max_val, rounds),
                'Min': round(min_val, rounds),
                'Var': round(var_val, rounds),
                'AVG': round(avg_val, rounds)
            })
    return pd.DataFrame(stats)


def get_stats_record(df, label_prefix, rounds):
    stats = []
    for col in df.columns:
        if col != 'Name':  # ignore text column
            stats.append({
                'Performance': f'{col} {label_prefix}',
                'Max': round(df[col].max(), rounds),
                'Min': round(df[col].min(), rounds),
                'Var': round(df[col].var(), rounds),
                'AVG': round(df[col].mean(), rounds)
            })
    return pd.DataFrame(stats)


def calculate_delta_vs_init(df):
    # Convert numeric columns
    for col in ['Max', 'Min', 'Var', 'AVG']:
        df[col] = df[col].astype(float)

    # Separate Init rows and others
    df_init = df[df['Versione'] == 'Init'].set_index('Performance')
    df_other = df[df['Versione'] != 'Init'].set_index('Performance')

    # Build a copy of df_init with same index as df_other
    df_init_expanded = df_other.copy()
    for idx in df_other.index:
        if 'Process' in idx:
            df_init_expanded.loc[idx] = df_init.loc['Process load Init']
        elif 'System' in idx:
            df_init_expanded.loc[idx] = df_init.loc['System load Init']

    # Create a new DataFrame to contain "value (delta)" strings
    df_formatted = df_other.copy()
    for col in ['Max', 'Min', 'Var', 'AVG']:
        final_values = df_other[col]
        ref_values = df_init_expanded[col]
        delta = final_values - ref_values
        df_formatted[col] = final_values.map(lambda x: f"{x:.2f}") + \
                             " (" + delta.map(lambda d: f"{d:+.2f}") + ")"

    # Reset index for display
    df_formatted = df_formatted.reset_index()
    return df_formatted


def calculate_delta_memory_vs_init(df):
    # Convert numeric columns
    for col in ['Max', 'Min', 'Var', 'AVG']:
        df[col] = df[col].astype(float)

    # Separate Init rows and others
    df_init = df[df['Versione'] == 'Init'].set_index('Performance')
    df_other = df[df['Versione'] != 'Init'].set_index('Performance')

    # Build a copy of df_init with same index as df_other
    df_init_expanded = df_other.copy()
    for idx in df_other.index:
        if 'Committed size' in idx:
            df_init_expanded.loc[idx] = df_init.loc['Committed size Init']
        elif 'Free size' in idx:
            df_init_expanded.loc[idx] = df_init.loc['Free size Init']
        elif 'Used size' in idx:
            df_init_expanded.loc[idx] = df_init.loc['Used size Init']

    # Create a new DataFrame to contain "value (delta)" strings
    df_formatted = df_other.copy()
    for col in ['Max', 'Min', 'Var', 'AVG']:
        final_values = df_other[col]
        ref_values = df_init_expanded[col]
        delta = final_values - ref_values
        df_formatted[col] = final_values.map(lambda x: f"{x:.2f}") + \
                             " (" + delta.map(lambda d: f"{d:+.2f}") + ")"

    df_formatted = df_formatted.reset_index()
    return df_formatted


def calculate_delta_record_vs_init(df):
    # Convert numeric columns
    for col in ['Max', 'Min', 'Var', 'AVG']:
        df[col] = df[col].astype(float)

    # Separate Init rows and others
    df_init = df[df['Versione'] == 'Init'].set_index('Performance')
    df_other = df[df['Versione'] != 'Init'].set_index('Performance')

    # Build a copy of df_init with same index as df_other
    df_init_expanded = df_other.copy()
    for idx in df_other.index:
        if 'Instance Count' in idx:
            df_init_expanded.loc[idx] = df_init.loc['Instance Count Init']
        elif 'Size (bytes)' in idx:
            df_init_expanded.loc[idx] = df_init.loc['Size (bytes) Init']

    # Create a new DataFrame to contain "value (delta)" strings
    df_formatted = df_other.copy()
    for col in ['Max', 'Min', 'Var', 'AVG']:
        final_values = df_other[col]
        ref_values = df_init_expanded[col]
        delta = final_values - ref_values
        df_formatted[col] = final_values.map(lambda x: f"{x:.2f}") + \
                             " (" + delta.map(lambda d: f"{d:+.2f}") + ")"

    df_formatted = df_formatted.reset_index()
    return df_formatted


def calculate_general_stats(dfInit, dfNoDependency, dfNoModular,
                            dfMemoryInit, dfMemoryNoDependency, dfMemoryNoModular,
                            dfRecordInit, dfRecordNoDependency, dfRecordNoModular,
                            project_name, rounds, module=""):

    # --- CPU
    df1 = get_stats(dfInit, init_label, rounds); df1["Versione"] = "Init"
    df2 = get_stats(dfNoModular, no_mod_label, rounds); df2["Versione"] = "NoModular"
    df3 = get_stats(dfNoDependency, no_dep_label, rounds); df3["Versione"] = "NoDependency"
    dfCPU = pd.concat([df1, df2, df3], ignore_index=True)
    dfCPU["project"] = project_name
    dfCPU["module"] = module
    dfDeltaCPU = calculate_delta_vs_init(dfCPU)

    # --- Memory
    df1 = get_stats_memory(dfMemoryInit, init_label, rounds); df1["Versione"] = "Init"
    df2 = get_stats_memory(dfMemoryNoModular, no_mod_label, rounds); df2["Versione"] = "NoModular"
    df3 = get_stats_memory(dfMemoryNoDependency, no_dep_label, rounds); df3["Versione"] = "NoDependency"
    dfMemory = pd.concat([df1, df2, df3], ignore_index=True)
    dfMemory["project"] = project_name
    dfMemory["module"] = module
    dfDeltaMemory = calculate_delta_memory_vs_init(dfMemory)

    # --- Recorded Objects
    df1 = get_stats_record(dfRecordInit, init_label, rounds); df1["Versione"] = "Init"
    df2 = get_stats_record(dfRecordNoModular, no_mod_label, rounds); df2["Versione"] = "NoModular"
    df3 = get_stats_record(dfRecordNoDependency, no_dep_label, rounds); df3["Versione"] = "NoDependency"
    dfRecord = pd.concat([df1, df2, df3], ignore_index=True)
    dfRecord["project"] = project_name
    dfRecord["module"] = module
    dfDeltaRecord = calculate_delta_record_vs_init(dfRecord)

    # Return both delta and raw data
    return dfDeltaCPU, dfDeltaMemory, dfDeltaRecord, dfCPU, dfMemory, dfRecord


def concatenate_df(df0, df1, df2=None, df3=None, drop_module=True):
    dfs = [df for df in [df1, df2, df3] if df is not None]

    if not dfs:
        return df0 if df0 is not None else pd.DataFrame()

    df_concat = pd.concat(dfs, ignore_index=True)

    # Clean the Performance column
    df_concat["Performance"] = df_concat["Performance"].str.replace("No Modular", "", regex=False)
    df_concat["Performance"] = df_concat["Performance"].str.replace("No Dependency", "", regex=False)
    df_concat["Performance"] = df_concat["Performance"].str.strip()

    # Remove the 'module' column if requested and if it exists
    if drop_module and 'module' in df_concat.columns:
        df_concat = df_concat.drop(columns=['module'])

    # Reorder columns: 'Versione' and 'project' at the beginning
    first_columns = ["Versione", "project"]
    other_columns = [col for col in df_concat.columns if col not in first_columns]
    df_concat = df_concat[first_columns + other_columns]

    # Optionally concatenate with df0
    if df0 is not None:
        df_concat = pd.concat([df0, df_concat])

    return df_concat




def write_excel(df, sheet_name, path_excel):
    import os
    file_exists = os.path.exists(path_excel)

    if file_exists:
        with pd.ExcelWriter(path_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(path_excel, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(path_excel)
    ws = wb[sheet_name]

    # Adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # Identify column indices
    col_project = col_module = col_version = col_performance = None
    for cell in ws[1]:
        if cell.value == "project":
            col_project = cell.column
        elif cell.value == "module":
            col_module = cell.column
        elif cell.value == "Versione":
            col_version = cell.column
        elif cell.value == "Performance":
            col_performance = cell.column

    # Styles
    fill_modular = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_dependency = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thick = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")

    last_row = ws.max_row
    last_col = ws.max_column

    data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2, max_row=last_row)]

    cpu_metrics = {"Process load", "System load"}
    mem_metrics = {"Committed size", "Free size", "Used size"}
    rec_metrics = {"Instance Count", "Size (bytes)"}

    # Background color and horizontal separators
    for i in range(len(data)):
        row_data = data[i]
        version = row_data[col_version - 1]
        performance = str(row_data[col_performance - 1]) if col_performance else ""

        # Background color
        fill = fill_modular if version == "NoModular" else fill_dependency if version == "NoDependency" else None
        if fill:
            for col in range(1, last_col + 1):
                ws.cell(row=i + 2, column=col).fill = fill

        # Thick separator
        if i < len(data) - 1:
            curr = data[i]
            nxt = data[i + 1]
            different_project = curr[col_project - 1] != nxt[col_project - 1]
            different_module = curr[col_module - 1] != nxt[col_module - 1] if col_module else False
            if different_project or different_module:
                for col in range(1, last_col + 1):
                    ws.cell(row=i + 2, column=col).border = Border(bottom=thick)

        # Thin separator
        if i < len(data) - 1:
            curr = data[i]
            next_ = data[i + 1]
            same_project = curr[col_project - 1] == next_[col_project - 1]
            same_module = curr[col_module - 1] == next_[col_module - 1] if col_module else True
            if same_project and same_module:
                perf_curr = str(curr[col_performance - 1])
                perf_next = str(next_[col_performance - 1])
                if (
                    (perf_curr in cpu_metrics and perf_next not in cpu_metrics) or
                    (perf_curr in mem_metrics and perf_next not in mem_metrics) or
                    (perf_curr in rec_metrics and perf_next not in rec_metrics)
                ):
                    for col in range(1, last_col + 1):
                        ws.cell(row=i + 2, column=col).border = Border(bottom=thin)

    # Color deltas inside parentheses
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
        row_perf = row[col_performance - 1].value if col_performance else ""
        for cell in row:
            val = str(cell.value) if cell.value else ""
            if "(+" in val or "(-" in val:
                try:
                    delta = val.split("(")[1].replace(")", "")
                    if "+0.00" in delta or "-0.00" in delta:
                        continue
                    is_positive = delta.startswith("+")
                except:
                    continue
                cell.font = Font(color="008000" if is_positive and row_perf == "Free size" else
                                 "FF0000" if is_positive else
                                 "FF0000" if row_perf == "Free size" else "008000")

    # Color Delta % Var and Delta % AVG based on sign (+/-)
    for col_idx in range(1, last_col + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header not in {"% Var", "% AVG"}:
            continue
        for row in range(2, last_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = str(cell.value).replace("\u200B", "").replace(" ", "") if cell.value else ""
            row_perf = ws.cell(row=row, column=col_performance).value if col_performance else ""
            if not val.endswith("%"):
                continue
            is_positive = val.startswith("+")
            cell.font = Font(color="008000" if is_positive and row_perf == "Free size" else
                             "FF0000" if is_positive else
                             "FF0000" if row_perf == "Free size" else "008000")

    # Vertical borders for all columns except between Var <-> % Var and AVG <-> % AVG
    columns = {ws.cell(row=1, column=c).value: c for c in range(1, last_col + 1)}
    col_var = columns.get("Var")
    col_delta_var = columns.get("% Var")
    col_avg = columns.get("AVG")
    col_delta_avg = columns.get("% AVG")

    for row in ws.iter_rows(min_row=1, max_row=last_row):
        for col_idx in range(1, last_col + 1):
            # Skip border only between Var and Delta % Var or AVG and Delta % AVG
            if (col_idx == col_var and col_idx + 1 == col_delta_var) or \
            (col_idx == col_avg and col_idx + 1 == col_delta_avg):
                continue

            cell = row[col_idx - 1]
            current_border = cell.border or Border()
            cell.border = Border(
                left=current_border.left,
                right=thin,
                top=current_border.top,
                bottom=current_border.bottom,
            )

    wb.save(path_excel)


def create_summary_from_Run1(path_excel, new_sheet):
    source_sheet = "Run1"  # With uppercase R

    columns_to_clear = ["Max", "Min", "Var", "AVG"]

    if not os.path.exists(path_excel):
        print(f"❌ File not found: {path_excel}")
        return

    wb = load_workbook(path_excel)

    if source_sheet not in wb.sheetnames:
        print(f"❌ The sheet '{source_sheet}' does not exist.")
        print(f"➡️ Available sheets: {wb.sheetnames}")
        return

    # If the new sheet already exists, delete it to avoid conflicts
    if new_sheet in wb.sheetnames:
        del wb[new_sheet]

    ws_orig = wb[source_sheet]
    ws_new = wb.copy_worksheet(ws_orig)
    ws_new.title = new_sheet

    # Find the columns to clear
    header = [str(cell.value).strip() for cell in ws_new[1]]
    col_indices = [i + 1 for i, col in enumerate(header) if col in columns_to_clear]

    # Clear only those columns
    for row in ws_new.iter_rows(min_row=2, max_row=ws_new.max_row):
        for col_idx in col_indices:
            row[col_idx - 1].value = None  # Clear the cell

    wb.save(path_excel)
    print(f"✅ Created sheet '{new_sheet}' with columns {columns_to_clear} cleared.")


def create_summary_with_improvements(path_excel):
    summary_sheet = "Improvement"
    create_summary_from_Run1(path_excel, summary_sheet)

    target_columns = ["Max", "Min", "Var", "AVG"]

    if not os.path.exists(path_excel):
        print(f"❌ File not found: {path_excel}")
        return

    wb = load_workbook(path_excel)
    ws_summary = wb[summary_sheet]
    header = [str(cell.value).strip() for cell in ws_summary[1]]
    col_indices = {col: i for i, col in enumerate(header) if col in target_columns}
    run_sheets = [name for name in wb.sheetnames if name.startswith("Run") and name != summary_sheet]

    for row_idx in range(2, ws_summary.max_row + 1):
        performance_cell = ws_summary.cell(row=row_idx, column=4).value  # Column D: "Performance"
        is_free_size_row = isinstance(performance_cell, str) and performance_cell.strip() == "Free size"

        for col_name, col_idx in col_indices.items():
            improved = 0
            total = 0

            for run_sheet_name in run_sheets:
                ws_run = wb[run_sheet_name]
                value = ws_run.cell(row=row_idx, column=col_idx + 1).value

                if value is None or not isinstance(value, str) or '(' not in value:
                    continue

                delta_raw = value.split('(')[1].replace(')', '').strip()
                try:
                    delta = float(delta_raw)
                except ValueError:
                    continue

                # Exclude ±0.0
                if abs(delta) == 0.0:
                    total += 1
                    continue

                total += 1
                if is_free_size_row:
                    if delta > 0.0:
                        improved += 1
                else:
                    if delta < 0.0:
                        improved += 1

            if total > 0:
                fraction = f"{improved}/{total}"
                cell = ws_summary.cell(row=row_idx, column=col_idx + 1)
                cell.value = fraction

                # Remove red/green font color
                font = cell.font
                if font.color and font.color.type == "rgb" and font.color.rgb:
                    color_code = font.color.rgb[-6:].upper()
                    if color_code in {"FF0000", "00B050", "008000", "00FF00"}:
                        cell.font = Font(
                            name=font.name,
                            size=font.size,
                            bold=font.bold,
                            italic=font.italic,
                            vertAlign=font.vertAlign,
                            underline=font.underline,
                            strike=font.strike,
                            color="000000"
                        )

    wb.save(path_excel)
    print("✅ Summary updated with special logic for 'Free size' (case-sensitive).")


def create_summary_with_regressions(path_excel):
    summary_sheet = "Regression"
    create_summary_from_Run1(path_excel, new_sheet=summary_sheet)

    target_columns = ["Max", "Min", "Var", "AVG"]

    if not os.path.exists(path_excel):
        print(f"❌ File not found: {path_excel}")
        return

    wb = load_workbook(path_excel)
    ws_summary = wb[summary_sheet]
    header = [str(cell.value).strip() for cell in ws_summary[1]]
    col_indices = {col: i for i, col in enumerate(header) if col in target_columns}
    run_sheets = [name for name in wb.sheetnames if name.startswith("Run") and name != summary_sheet]

    for row_idx in range(2, ws_summary.max_row + 1):
        performance_cell = ws_summary.cell(row=row_idx, column=4).value  # Column D: "Performance"
        is_free_size_row = isinstance(performance_cell, str) and performance_cell.strip() == "Free size"

        for col_name, col_idx in col_indices.items():
            regressed = 0
            total = 0

            for run_sheet_name in run_sheets:
                ws_run = wb[run_sheet_name]
                value = ws_run.cell(row=row_idx, column=col_idx + 1).value

                if value is None or not isinstance(value, str) or '(' not in value:
                    continue

                delta_raw = value.split('(')[1].replace(')', '').strip()
                try:
                    delta = float(delta_raw)
                except ValueError:
                    continue

                # Always count in total (exclude only ±0.0)
                if abs(delta) == 0.0:
                    total += 1
                    continue

                total += 1

                if is_free_size_row:
                    if delta < 0.0:  # Free size → regression if NEGATIVE
                        regressed += 1
                else:
                    if delta > 0.0:  # Others → regression if POSITIVE
                        regressed += 1

            if total > 0:
                fraction = f"{regressed}/{total}"
                cell = ws_summary.cell(row=row_idx, column=col_idx + 1)
                cell.value = fraction

                # Remove red/green font color
                font = cell.font
                if font.color and font.color.type == "rgb" and font.color.rgb:
                    color_code = font.color.rgb[-6:].upper()
                    if color_code in {"FF0000", "00B050", "008000", "00FF00"}:
                        cell.font = Font(
                            name=font.name,
                            size=font.size,
                            bold=font.bold,
                            italic=font.italic,
                            vertAlign=font.vertAlign,
                            underline=font.underline,
                            strike=font.strike,
                            color="000000"
                        )

    wb.save(path_excel)
    print("✅ 'Regression' sheet updated: +delta means regression (except 'Free size' where -delta applies).")

def move_column(columns, target_col, col_to_move):
    if col_to_move in columns and target_col in columns:
        columns.remove(col_to_move)
        idx = columns.index(target_col)
        columns.insert(idx + 1, col_to_move)
    return columns


def calculate_average(df):
    """
    Calculate the average of metrics for each logical combination in the DataFrame,
    dynamically adapting to the presence of the 'module' column.
    Preserves the original appearance order.
    """
    metric_columns = ['Max', 'Min', 'Var', 'AVG']

    # Determine available groups based on present columns
    possible_groups = ['project', 'module', 'Performance', 'Versione']
    groups = [col for col in possible_groups if col in df.columns]

    # Compute average for actual groups
    df_avg = df.groupby(groups, dropna=False)[metric_columns].mean().reset_index()

    # Preserve the original order
    order = df[groups].drop_duplicates().reset_index(drop=True)
    df_avg = order.merge(df_avg, on=groups, how='left')

    return df_avg


def format_delta_vs_init_unified(df):
    df = df.copy()
    for col in ['Max', 'Min', 'Var', 'AVG']:
        df[col] = df[col].astype(float)

    # Dynamic index based on available columns
    idx_cols = ['project', 'Performance']
    if 'module' in df.columns:
        idx_cols.insert(1, 'module')  # Insert 'module' after 'project'

    # Split between init and others
    df_init = df[df['Versione'] == 'Init'].set_index(idx_cols)
    df_other = df[df['Versione'] != 'Init'].set_index(idx_cols)

    # Build DataFrame with corresponding Init values
    df_init_expanded = df_other.copy()
    for idx in df_other.index:
        perf = idx[-1]
        perf_init = perf + " Init"
        idx_init = idx[:-1] + (perf_init,)
        if idx_init in df_init.index:
            df_init_expanded.loc[idx, ['Max', 'Min', 'Var', 'AVG']] = df_init.loc[idx_init, ['Max', 'Min', 'Var', 'AVG']].values
        else:
            print(f"⚠️ Warning: {idx_init} not found in df_init")

    # Format deltas
    df_formatted = df_other.copy()
    for col in ['Max', 'Min', 'Var', 'AVG']:
        final_values = df_other[col]
        ref_values = df_init_expanded[col]
        delta = final_values - ref_values
        df_formatted[col] = final_values.map(lambda x: f"{x:.2f}") + \
                             " (" + delta.map(lambda d: f"{d:+.2f}") + ")"

    # Calculate percentage differences (only for AVG and Var)
    for col in ['Var', 'AVG']:
        final_values = df_other[col]
        ref_values = df_init_expanded[col]
        with np.errstate(divide='ignore', invalid='ignore'):
            perc_diff = np.where(ref_values != 0, (final_values - ref_values) / np.abs(ref_values) * 100, np.nan)
        perc_diff_str = [f"{v:+.2f}%\u200B" if not np.isnan(v) else "NaN" for v in perc_diff]

        df_formatted[f"% {col}"] = perc_diff_str

    # Reorder columns: % Var after Var, % AVG after AVG
    columns = list(df_formatted.columns)
    columns = move_column(columns, 'Var', '% Var')
    columns = move_column(columns, 'AVG', '% AVG')

    df_formatted = df_formatted[columns]
    df_formatted = df_formatted.reset_index()

    return df_formatted


def move_sheet_before(path_excel, sheet_to_move, destination_sheet):
    wb = load_workbook(path_excel)

    if sheet_to_move in wb.sheetnames and destination_sheet in wb.sheetnames:
        ws_from = wb[sheet_to_move]
        ws_to = wb[destination_sheet]
        idx = wb.worksheets.index(ws_to)

        # Remove and reinsert at the desired index
        wb._sheets.remove(ws_from)
        wb._sheets.insert(idx, ws_from)

        wb.save(path_excel)
        print(f"✅ Moved '{sheet_to_move}' before '{destination_sheet}' in {path_excel}")
    else:
        print(f"⚠️ One of the sheets '{sheet_to_move}' or '{destination_sheet}' does not exist in {path_excel}")


def calculate_module_average(df):
    metric_columns = ['Max', 'Min', 'Var', 'AVG']
    groups = ['project', 'Versione', 'Performance']

    # Remove 'module' column if present
    if 'module' in df.columns:
        df = df.drop(columns=['module'])

    # Preserve original order of combinations
    order = df[groups].drop_duplicates().reset_index(drop=True)

    # Calculate average of metrics
    df_avg = df.groupby(groups, dropna=False)[metric_columns].mean().reset_index()

    # Merge to keep original order
    df_avg = order.merge(df_avg, on=groups, how='left')

    return df_avg


# === FINAL EXECUTION LOGIC ===

df_concat_general_stats = None
df_concat_general_stats_modules = {}
if os.path.exists(output_file):
    os.remove(output_file)

for k in range(runs):
    df_concat = None
    df_concat_modules = {}

    for i, name in enumerate(projects):
        print("----------------------------")
        print(f"Project performance: {name} — execution {k+1} of {runs}")

        if name in MODULES_PROJECTS:
            modules = MODULES_PROJECTS[name]
            df_module_concat = None

            for module in modules:
                print(f"\n➤ Module: {module}")
                (
                    dfInit, dfNoModular, dfNoDependency,
                    dfMemoryInit, dfMemoryNoModular, dfMemoryNoDependency,
                    dfRecordInit, dfRecordNoModular, dfRecordNoDependency
                ) = load_csv_comparison(name, module, str(k + 1), BASE_DIR)
                if dfInit is not None:
                    _, _, _, dfCPU, dfMemory, dfRecord = calculate_general_stats(
                        dfInit, dfNoDependency, dfNoModular,
                        dfMemoryInit, dfMemoryNoDependency, dfMemoryNoModular,
                        dfRecordInit, dfRecordNoDependency, dfRecordNoModular,
                        name, rounds, module
                    )

                    for df in [dfCPU, dfMemory, dfRecord]:
                        df["Run"] = k + 1

                    df_module_concat = concatenate_df(df_module_concat, dfCPU, dfMemory, dfRecord, drop_module=False)

            # Compute average and delta
            df_module_avg = calculate_module_average(df_module_concat)
            df_delta = format_delta_vs_init_unified(df_module_avg)

            df_concat = concatenate_df(df_concat, df_delta)
            df_concat_general_stats = concatenate_df(df_concat_general_stats, df_module_avg)

            # Second Excel: direct delta from concat
            df_delta_concat = format_delta_vs_init_unified(df_module_concat)
            if name not in df_concat_modules or df_concat_modules[name] is None:
                df_concat_modules[name] = df_delta_concat
                df_concat_general_stats_modules[name] = df_module_concat
            else:
                df_concat_modules[name] = concatenate_df(df_concat_modules[name], df_delta_concat)
                df_concat_general_stats_modules[name] = concatenate_df(df_concat_general_stats_modules[name], df_module_concat)

        else:
            print(f"\n➤ Project without modules")

            module = ""

            (
                dfInit, dfNoModular, dfNoDependency,
                dfMemoryInit, dfMemoryNoModular, dfMemoryNoDependency,
                dfRecordInit, dfRecordNoModular, dfRecordNoDependency
            ) = load_csv_comparison(name, module, str(k+1), BASE_DIR)

            dfDeltaCPU, dfDeltaMemory, dfDeltaRecord, dfCPU, dfMemory, dfRecord = calculate_general_stats(
                dfInit, dfNoDependency, dfNoModular,
                dfMemoryInit, dfMemoryNoDependency, dfMemoryNoModular,
                dfRecordInit, dfRecordNoDependency, dfRecordNoModular,
                name, rounds, ""
            )

            for df in [dfCPU, dfMemory, dfRecord]:
                df["Run"] = k + 1

            df_concat = concatenate_df(df_concat, dfDeltaCPU, dfDeltaMemory, dfDeltaRecord)
            df_concat_general_stats = concatenate_df(df_concat_general_stats, dfCPU, dfMemory, dfRecord)

    # Save current run to Excel
    df_concat = df_concat.drop(columns=["% Var", "% AVG"], errors="ignore")
    write_excel(df_concat, sheet_name=f"Run{k+1}", path_excel=output_file)

    for name in df_concat_modules:
        write_excel(df_concat_modules[name], sheet_name=f"Run{k+1}", path_excel=output_file.replace(".xlsx", f"_{name}.xlsx"))

# Calculate and save final average
df_avg = calculate_average(df_concat_general_stats)
df_avg_delta = format_delta_vs_init_unified(df_avg)
write_excel(df_avg_delta, sheet_name="FinalAverage", path_excel=output_file)
move_sheet_before(output_file, "FinalAverage", "Run1")

# Final average per modular project
for name, df_stat in df_concat_general_stats_modules.items():
    print(f"📊 Writing final average for project modules: {name}")
    df_avg_modules = calculate_module_average(df_stat)
    df_avg_modules_delta = format_delta_vs_init_unified(df_avg_modules)
    write_excel(df_avg_modules_delta, sheet_name="FinalAverage", path_excel=output_file.replace(".xlsx", f"_{name}.xlsx"))

create_summary_with_improvements(output_file)
create_summary_with_regressions(output_file)
